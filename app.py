# app_etichette.py — Generatore Etichette Poste Delivery Business Express
# Versione Streamlit — layout 113.96 × 104.01 mm — Zebra ZT410/ZT411 203 dpi

import streamlit as st
import os, tempfile, hashlib, re, io

from reportlab.lib.units import mm as rmm
from reportlab.lib import colors
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.units import mm
from PIL import Image as PILImage, ImageDraw

st.set_page_config(
    page_title="Etichette Poste Business Express",
    page_icon="🏷️",
    layout="wide",
)

st.markdown("""
<style>
    .header-bar {
        background: #111;
        padding: 18px 24px 12px;
        border-radius: 8px;
        margin-bottom: 20px;
    }
    .header-title { color: white; font-size: 1.3rem; font-weight: 700; margin: 0; }
    .header-sub   { color: #aaa;   font-size: 0.78rem; margin: 4px 0 0; }
    .section-label {
        font-size: 0.8rem; font-weight: 700;
        color: #555; text-transform: uppercase;
        letter-spacing: 0.05em; margin-bottom: 4px;
    }
</style>
<div class="header-bar">
    <p class="header-title">🏷️ postedeliverybusiness express — Generatore Etichette</p>
    <p class="header-sub">113.96 × 104.01 mm &nbsp;|&nbsp; Zebra ZT410 / ZT411 &nbsp;|&nbsp; 203 dpi</p>
</div>
""", unsafe_allow_html=True)

PW = 113.96
PH = 104.01

# ─── BARCODE Code128-B ────────────────────────────────────────────────────────
_C128_STR = (" !\"#$%&'()*+,-./0123456789:;<=>?@ABCDEFGHIJKLMNOPQRSTUVWXYZ"
             "[\\]^_`abcdefghijklmnopqrstuvwxyz{|}~")
_C128_MAP = {ch: i for i, ch in enumerate(_C128_STR)}
_C128_PAT = [
    "11011001100","11001101100","11001100110","10010011000","10010001100",
    "10001001100","10011001000","10011000100","10001100100","11001001000",
    "11001000100","11000100100","10110011100","10011011100","10011001110",
    "10111001100","10011101100","10011100110","11001110010","11001011100",
    "11001001110","11011100100","11001110100","11101101110","11101001100",
    "11100101100","11100100110","11101100100","11100110100","11100110010",
    "11011011000","11011000110","11000110110","10100011000","10001011000",
    "10001000110","10110001000","10001101000","10001100010","11010001000",
    "11000101000","11000100010","10110111000","10110001110","10001101110",
    "10111011000","10111000110","10001110110","11101110110","11010001110",
    "11000101110","11011101000","11011100010","11011101110","11101011000",
    "11101000110","11100010110","11011010000","11011001010","11011110010",
    "11011011110","11001000010","11110001010","10100110000","10100001100",
    "10010110000","10010000110","10000101100","10000100110","10110010000",
    "10110000100","10011010000","10011000010","10000110100","10000110010",
    "11000010010","11001010000","11110111010","11000010100","10001111010",
    "10100111100","10010111100","10010011110","10111100100","10011110100",
    "10011110010","11110100100","11110010100","11110010010","11011011110",
    "11011110110","11110110110","10101111000","10100011110","10001011110",
    "10111101000","10111100010","11110101000","11110100010","10111011110",
    "10111101110","11101011110","11110101110","11010000100","11010010000",
    "11010011100",
]
_START_B = 104
_STOP_PAT = "1100011101011"

def _bits(text):
    codes = [_START_B]; chk = _START_B
    for i, ch in enumerate(text):
        v = _C128_MAP.get(ch, 16); codes.append(v); chk += v*(i+1)
    codes.append(chk % 103)
    b = _C128_PAT[_START_B]
    for v in codes[1:]:
        if v < len(_C128_PAT): b += _C128_PAT[v]
    return b + _STOP_PAT + "11"

def barcode_img(text, h=60, vertical=False, mod=4) -> PILImage.Image:
    MOD  = mod
    bits = _bits(text)
    W    = len(bits) * MOD + MOD * 4
    img  = PILImage.new("L", (W, h), 255)
    draw = ImageDraw.Draw(img)
    x = MOD * 2
    for bit in bits:
        if bit == "1":
            draw.rectangle([x, 0, x + MOD - 1, h - 1], fill=0)
        x += MOD
    if vertical:
        img = img.rotate(90, expand=True)
    return img

def make_qr(text: str, size: int = 100) -> PILImage.Image:
    try:
        import qrcode
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M,
                           box_size=10, border=2)
        qr.add_data(text); qr.make(fit=True)
        return qr.make_image(fill_color="black", back_color="white")\
                 .convert("L").resize((size, size), PILImage.NEAREST)
    except ImportError:
        pass
    N = 25; cell = max(2, size // N); tot = N * cell
    img = PILImage.new("L", (tot, tot), 255)
    d = ImageDraw.Draw(img)
    def fp(ox, oy):
        d.rectangle([ox, oy, ox+7*cell-1, oy+7*cell-1], fill=0)
        d.rectangle([ox+cell, oy+cell, ox+6*cell-1, oy+6*cell-1], fill=255)
        d.rectangle([ox+2*cell, oy+2*cell, ox+5*cell-1, oy+5*cell-1], fill=0)
    fp(2,2); fp(tot-7*cell-2,2); fp(2,tot-7*cell-2)
    bits2 = bin(int(hashlib.sha256(text.encode()).hexdigest(),16))[2:].zfill(256)
    idx = 0
    for r in range(N):
        for col in range(N):
            skip = (r<9 and col<9) or (r<9 and col>=N-8) or (r>=N-8 and col<9)
            if not skip:
                if bits2[idx % len(bits2)] == "1":
                    d.rectangle([col*cell+2,r*cell+2,col*cell+cell+1,r*cell+cell+1],fill=0)
                idx += 1
    return img.resize((size, size), PILImage.NEAREST)

_ECC200_SIZES = [
    (10,10,1,1,3,5),(12,12,1,1,5,7),(14,14,1,1,8,10),(16,16,1,1,12,12),
    (18,18,1,1,18,14),(20,20,1,1,22,18),(22,22,1,1,30,20),(24,24,1,1,36,24),
    (26,26,1,1,44,28),(32,32,2,2,62,36),(36,36,2,2,86,42),(40,40,2,2,114,48),
    (44,44,2,2,144,56),(48,48,2,2,174,68),(52,52,2,2,204,84),(64,64,4,4,280,112),
    (72,72,4,4,368,144),(80,80,4,4,456,192),(88,88,4,4,576,224),(96,96,4,4,696,272),
    (104,104,4,4,816,336),(120,120,6,6,1050,408),(132,132,6,6,1304,496),(144,144,6,6,1558,620),
]

def _make_dm_matrix(text: str):
    from reportlab.graphics.barcode.ecc200datamatrix import ECC200DataMatrix
    t = text.upper()
    for row, col, rr, cr, cw_data, cw_ecc in _ECC200_SIZES:
        try:
            obj = ECC200DataMatrix(t)
            obj.row_modules=row; obj.col_modules=col
            obj.row_regions=rr;  obj.col_regions=cr
            obj.cw_data=cw_data; obj.cw_ecc=cw_ecc
            obj.row_usable_modules=row-rr*2
            obj.col_usable_modules=col-cr*2
            obj.validate(); matrix=obj.encode()
            return matrix
        except Exception:
            continue
    raise Exception("DataMatrix: testo troppo lungo anche per 144x144 ECC200.")

def make_datamatrix(text: str, size: int = 100) -> PILImage.Image:
    try:
        from pylibdmtx.pylibdmtx import encode as _dm_enc
        raw = _dm_enc(text.encode("utf-8"))
        img = PILImage.frombytes("RGB",(raw.width,raw.height),raw.pixels)
        return img.convert("L").resize((size,size),PILImage.NEAREST)
    except Exception: pass
    try:
        import segno as _sg, io as _io
        dm = _sg.make(text, kind="dmtx")
        buf = _io.BytesIO()
        dm.save(buf, kind="png", scale=max(1,size//20), border=1)
        buf.seek(0)
        return PILImage.open(buf).convert("L").resize((size,size),PILImage.NEAREST)
    except Exception: pass
    matrix = _make_dm_matrix(text)
    nrows = len(matrix); ncols = len(matrix[0]) if nrows else 0
    quiet = 2; cell = max(2, size // max(nrows, ncols))
    W2 = (ncols + 2*quiet)*cell; H2 = (nrows + 2*quiet)*cell
    img = PILImage.new("L",(W2,H2),255)
    d = ImageDraw.Draw(img)
    for r, row in enumerate(matrix):
        for col2, val in enumerate(row):
            if val:
                x0=(col2+quiet)*cell; y0=(r+quiet)*cell
                d.rectangle([x0,y0,x0+cell-1,y0+cell-1],fill=0)
    return img.resize((size,size),PILImage.NEAREST)

def _tmp(img):
    p = tempfile.mktemp(suffix=".png"); img.save(p); return p

def _draw_cap_text(c, text, x_mm, baseline_y_mm, font_size_pt=28):
    c.setFont("Helvetica-Bold", font_size_pt)
    natural_w = c.stringWidth(text)
    target_w_pt = 22.5 * 2.8346
    scale_x = min(1.0, target_w_pt / max(natural_w, 1))
    c.saveState()
    c.setFillColor(colors.white)
    c.transform(scale_x, 0, 0, 1.0, x_mm*rmm, baseline_y_mm*rmm)
    c.drawString(0, 0, text)
    c.restoreState()

# ─── GENERATORE PDF COMPLETO ──────────────────────────────────────────────────
def genera_pdf(dati: dict, output_path: str, accessorio: str = "", extra_code: str = ""):
    W = PW * rmm; H = PH * rmm
    c = rl_canvas.Canvas(output_path, pagesize=(W, H))

    codice_base = dati.get("codice_spedizione","").strip()
    cap         = dati.get("cap","").strip()
    cap_code    = (cap + "0"*9)[:9]

    accessorio  = accessorio.strip().upper()
    extra_code  = extra_code.strip().upper()
    ha_extra    = bool(accessorio and accessorio != "APP" and extra_code)
    codice      = accessorio + codice_base if accessorio else codice_base

    c.setFillColor(colors.white)
    c.rect(0, 0, W, H, fill=1, stroke=0)

    # 1. Banda nera CAP
    c.setFillColor(colors.black)
    c.rect(5.19*rmm, 87.98*rmm, 25.50*rmm, 12.42*rmm, fill=1, stroke=0)
    _draw_cap_text(c, f"CAP {cap}", 6, 88.7)

    # DataMatrix string
    _cap_dest = cap; _citta_d = ""; _prov_d = ""; _cap_mit = ""; _citta_m = ""; _prov_m = ""
    cap_citta_d = dati.get("dest_cap_citta","")
    if cap_citta_d:
        parts = cap_citta_d.split()
        if parts: _cap_dest = parts[0]
        rest = " ".join(parts[1:]).rstrip(")").lstrip("(")
        if "(" in rest:
            _citta_d = rest[:rest.index("(")].strip(); _prov_d = rest[rest.index("(")+1:].strip()
        else: _citta_d = rest
    cap_citta_m = dati.get("mit_cap_citta","")
    if cap_citta_m:
        parts = cap_citta_m.split()
        if parts: _cap_mit = parts[0]
        rest = " ".join(parts[1:]).rstrip(")").lstrip("(")
        if "(" in rest:
            _citta_m = rest[:rest.index("(")].strip(); _prov_m = rest[rest.index("(")+1:].strip()
        else: _citta_m = rest

    dm_big_text = (
        f"{codice}|{_cap_dest}||{dati.get('dest_nome','').upper()}|-|"
        f"{dati.get('dest_indirizzo','').upper()}|{_cap_dest}|{_citta_d.upper()}|{_prov_d.upper()}|ITA|"
        f"{dati.get('colli','')}|{dati.get('mit_nome','').upper()}||"
        f"{dati.get('mit_indirizzo','').upper()}|{_cap_mit}|{_citta_m.upper()}|{_prov_m.upper()}|ITA|PBE||||"
        f"{dati.get('dest_tel','')}|||{dati.get('peso','')}|"
    )

    # 3. QR piccolo
    qr_s = make_qr(codice, size=100); qrsp = _tmp(qr_s)
    c.drawImage(qrsp, 64.98*rmm, 87.9*rmm, width=13.23*rmm, height=13.23*rmm, preserveAspectRatio=True)
    os.unlink(qrsp)

    # 4. DataMatrix grande
    dm_b = make_datamatrix(dm_big_text, size=250); dmbp = _tmp(dm_b)
    c.drawImage(dmbp, 81.18*rmm, 75.84*rmm, width=26.46*rmm, height=26.46*rmm, preserveAspectRatio=True)
    os.unlink(dmbp)

    # 5. Barcode sinistro LDV
    GS1_MIN = 0.191; GAP_BC = 2.0
    SX_YBOT=4.87; SX_YTOP=87.98; SX_AVAIL=SX_YTOP-SX_YBOT
    SX_X=2.01; SX_W=25.12; SX_TXT_X=29.5
    DX_YBOT=6.84; DX_YTOP=75.84; DX_AVAIL=DX_YTOP-DX_YBOT
    DX_X=94.23; DX_W=15.35

    def _n_bit(text): return len(_bits(text))
    def _split_heights(avail, n_main, n_extra, gap, gmin, ratio=0.70):
        main_h_min=gmin*n_main; extra_h_min=gmin*n_extra
        main_h=max(main_h_min, avail*ratio); extra_h=avail-gap-main_h
        if extra_h < extra_h_min:
            extra_h=extra_h_min; main_h=avail-gap-extra_h
        return round(main_h,2), round(extra_h,2)

    ldv_h = SX_AVAIL
    bcv = barcode_img(codice, h=400, vertical=True, mod=4); bcvp = _tmp(bcv)
    c.drawImage(bcvp, SX_X*rmm, SX_YBOT*rmm, width=SX_W*rmm, height=ldv_h*rmm, preserveAspectRatio=False)
    os.unlink(bcvp)
    c.saveState(); c.setFont("Helvetica",6)
    c.translate(SX_TXT_X*rmm, (SX_YBOT+ldv_h/2)*rmm); c.rotate(90)
    c.drawCentredString(0,0,codice_base); c.restoreState()

    # 6. Barcode destro CAP
    n_cap = _n_bit(cap_code)
    if ha_extra:
        n_ext_dx = _n_bit(extra_code)
        cap_h, ext_h_dx = _split_heights(DX_AVAIL, n_cap, n_ext_dx, GAP_BC, GS1_MIN, ratio=0.50)
        ext_ybot_dx = DX_YBOT + cap_h + GAP_BC
    else:
        cap_h = DX_AVAIL

    bcv2 = barcode_img(cap_code, h=400, vertical=True, mod=4); bcv2p = _tmp(bcv2)
    c.drawImage(bcv2p, DX_X*rmm, DX_YBOT*rmm, width=DX_W*rmm, height=cap_h*rmm, preserveAspectRatio=False)
    os.unlink(bcv2p)
    c.saveState(); c.setFont("Helvetica",6)
    c.translate((DX_X+DX_W+1.5)*rmm,(DX_YBOT+cap_h/2)*rmm); c.rotate(90)
    c.drawCentredString(0,0,cap_code); c.restoreState()

    if ha_extra:
        bcvE = barcode_img(extra_code, h=400, vertical=True, mod=4); bcvEp = _tmp(bcvE)
        c.drawImage(bcvEp, DX_X*rmm, ext_ybot_dx*rmm, width=DX_W*rmm, height=ext_h_dx*rmm, preserveAspectRatio=False)
        os.unlink(bcvEp)
        c.saveState(); c.setFont("Helvetica",6)
        c.translate((DX_X+DX_W+1.5)*rmm,(ext_ybot_dx+ext_h_dx/2)*rmm); c.rotate(90)
        c.drawCentredString(0,0,extra_code); c.restoreState()
        c.setFont("Helvetica-Bold",5); c.setFillColor(colors.black)
        c.drawCentredString((DX_X+DX_W/2)*rmm,(ext_ybot_dx+ext_h_dx+1.0)*rmm, accessorio)

    # 7. Box principale
    bx=32.0*rmm; bw=62.23*rmm; byt=70.49*rmm; byb=3.46*rmm; bh=byt-byb
    c.setStrokeColor(colors.black); c.setLineWidth(0.5)
    c.rect(bx,byb,bw,bh,fill=0,stroke=1)
    c.line(bx,56.02*rmm,bx+bw,56.02*rmm)
    c.line(bx,31.68*rmm,bx+bw,31.68*rmm)
    c.line(bx,27.09*rmm,bx+bw,27.09*rmm)

    # 8. Mittente
    c.setFont("Helvetica-Bold",6.8); c.setFillColor(colors.black)
    c.drawString(33.06*rmm,67.5*rmm,"Mittente")
    c.setFont("Helvetica",6.8)
    c.drawString(43.0*rmm,67.5*rmm,"TEL:")
    c.drawString(50.0*rmm,67.5*rmm,dati.get("mit_tel",""))
    c.setFont("Helvetica",6.0)
    c.drawString(33.06*rmm,64.3*rmm,dati.get("mit_nome",""))
    c.drawString(33.06*rmm,61.8*rmm,dati.get("mit_indirizzo",""))
    c.drawString(33.06*rmm,59.3*rmm,dati.get("mit_cap_citta",""))

    # 9. Destinatario
    c.setFont("Helvetica-Bold",6.8)
    c.drawString(33.06*rmm,53*rmm,"Destinatario")
    c.setFont("Helvetica",6.8)
    c.drawString(49.0*rmm,53*rmm,"TEL:")
    c.drawString(56.0*rmm,53*rmm,dati.get("dest_tel",""))
    c.setFont("Helvetica",6.0)
    c.drawString(33.06*rmm,49.5*rmm,dati.get("dest_nome",""))
    c.drawString(33.06*rmm,42.5*rmm,dati.get("dest_indirizzo",""))
    c.drawString(33.06*rmm,40.0*rmm,dati.get("dest_cap_citta",""))

    # 10. Lettera di vettura
    c.setFont("Helvetica-Bold",5.2)
    c.drawString(32.79*rmm,29.8*rmm,"Lettera di vettura")
    c.drawString(62.0*rmm,29.8*rmm,"PESO(KG)")
    c.drawString(74.4*rmm,29.8*rmm,"COLLI")
    c.drawString(82.8*rmm,29.8*rmm,"HxLxP(CM)")
    c.setFont("Helvetica",4.6)
    c.drawString(32.79*rmm,28.2*rmm,codice)
    c.setFont("Helvetica",6.2)
    c.drawString(63.0*rmm,28.2*rmm,dati.get("peso",""))
    c.drawString(75.0*rmm,28.2*rmm,dati.get("colli",""))
    c.drawString(82.8*rmm,28.2*rmm,dati.get("dimensioni",""))
    if accessorio:
        c.setFont("Helvetica-Bold",5.2)
        acc_label = f"Accessorio: {accessorio}"
        if ha_extra: acc_label += f"  —  Rif.: {extra_code}"
        c.drawString(32.79*rmm,25.5*rmm,acc_label)

    c.showPage(); c.save()


# ─── GENERATORE PDF SOLO BARCODE ─────────────────────────────────────────────
def genera_pdf_solo_barcode(codice: str, cap: str, output_path: str,
                             accessorio: str = "", extra_code: str = ""):
    accessorio    = accessorio.strip().upper()
    extra_code    = extra_code.strip().upper()
    ha_extra      = bool(accessorio and accessorio != "APP" and extra_code)
    codice_finale = accessorio + codice if accessorio else codice

    W = PW * rmm; H = PH * rmm
    c = rl_canvas.Canvas(output_path, pagesize=(W, H))
    cap_code = (cap + "0"*9)[:9]

    c.setFillColor(colors.white); c.rect(0,0,W,H,fill=1,stroke=0)

    MARGIN=4.0; MARGIN_SIDES=PW*(12.0/PW)
    bc_w_mm=PW-MARGIN_SIDES; cap_w_mm=bc_w_mm/2
    x_ldv=(PW-bc_w_mm)/2; x_cap=(PW-cap_w_mm)/2
    txt_h=PH*(3.5/PH); bc_h_mm=18.0; cap_h_mm=14.0; extra_h=14.0
    gap_inner=6.0; gap_outer=12.0

    if ha_extra:
        block_h = extra_h+txt_h+gap_inner+bc_h_mm+txt_h+gap_outer+cap_h_mm+txt_h
    else:
        block_h = bc_h_mm+txt_h+gap_outer+cap_h_mm+txt_h

    y_top = (PH + block_h) / 2

    if ha_extra:
        y_bot_extra=y_top-extra_h; y_txt_extra=y_bot_extra-txt_h
        y_bot_ldv=y_txt_extra-gap_inner-bc_h_mm; y_txt_ldv=y_bot_ldv-txt_h
    else:
        y_bot_ldv=y_top-bc_h_mm; y_txt_ldv=y_bot_ldv-txt_h

    y_bot_cap=y_txt_ldv-gap_outer-cap_h_mm; y_txt_cap=y_bot_cap-txt_h
    c.setFillColor(colors.black)

    if ha_extra:
        bc_ext=barcode_img(extra_code,h=400,vertical=False,mod=4); p_ext=_tmp(bc_ext)
        c.drawImage(p_ext,x_ldv*rmm,y_bot_extra*rmm,width=bc_w_mm*rmm,height=extra_h*rmm,preserveAspectRatio=False)
        os.unlink(p_ext)
        c.setFont("Helvetica",5); c.drawCentredString((PW/2)*rmm,y_txt_extra*rmm,f"{accessorio}  {extra_code}")

    bc_ldv=barcode_img(codice_finale,h=400,vertical=False,mod=4); p_ldv=_tmp(bc_ldv)
    c.drawImage(p_ldv,x_ldv*rmm,y_bot_ldv*rmm,width=bc_w_mm*rmm,height=bc_h_mm*rmm,preserveAspectRatio=False)
    os.unlink(p_ldv)
    c.setFont("Helvetica",5); c.setFillColor(colors.black)
    c.drawCentredString((PW/2)*rmm,y_txt_ldv*rmm,codice_finale)

    bc_cap=barcode_img(cap_code,h=400,vertical=False,mod=4); p_cap=_tmp(bc_cap)
    c.drawImage(p_cap,x_cap*rmm,y_bot_cap*rmm,width=cap_w_mm*rmm,height=cap_h_mm*rmm,preserveAspectRatio=False)
    os.unlink(p_cap)
    c.setFont("Helvetica",5); c.drawCentredString((PW/2)*rmm,y_txt_cap*rmm,cap)

    c.showPage(); c.save()


# ─── IMPORTA DA EXCEL (OneTracking) ──────────────────────────────────────────
def leggi_righe_xls(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return [[cell.value for cell in row] for row in ws.iter_rows()]
    with open(path, "rb") as fh:
        header = fh.read(6)
    if header[:5] in (b"<?xml", b"<Work") or header[:2] != b'\xd0\xcf':
        import xml.etree.ElementTree as ET
        tree = ET.parse(path); root = tree.getroot()
        ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
        righe_out = []
        for ws_el in root.findall(".//ss:Worksheet", ns):
            for row_el in ws_el.findall(".//ss:Row", ns):
                riga = []
                for cell_el in row_el.findall("ss:Cell", ns):
                    data_el = cell_el.find("ss:Data", ns)
                    riga.append(data_el.text if data_el is not None else None)
                righe_out.append(riga)
        return righe_out
    try:
        import xlrd
        wb = xlrd.open_workbook(path); ws = wb.sheet_by_index(0)
        return [ws.row_values(r) for r in range(ws.nrows)]
    except ImportError:
        raise RuntimeError("Installa xlrd per file .xls binario:\n  pip install xlrd")

def parse_indirizzo(raw):
    if not raw: return "","","",""
    raw = str(raw).strip()
    m = re.match(r'^(.*?),?\s+(\d{5})\s+(.+?)(?:\s+\(([A-Z]{2})\))?$', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2), m.group(3).strip(), (m.group(4) or "")
    return raw,"","",""

def importa_da_excel(file_bytes, ext) -> dict:
    tmp = tempfile.mktemp(suffix=ext)
    with open(tmp,"wb") as f: f.write(file_bytes)
    try:
        righe = leggi_righe_xls(tmp)
    finally:
        os.unlink(tmp)

    dati = {k:"" for k in ["codice_spedizione","cap","mit_nome","mit_indirizzo",
                             "mit_cap_citta","mit_tel","dest_nome","dest_indirizzo",
                             "dest_cap_citta","dest_tel","peso","colli","dimensioni"]}

    intestazione_lv = False
    for riga in righe:
        cells = [str(c).strip() if c is not None else "" for c in riga]
        testo_riga = " ".join(cells).strip().upper()
        if "LETTERA DI VETTURA" in testo_riga or "CODICE SPEDIZIONE" in testo_riga:
            intestazione_lv = True; continue
        if intestazione_lv:
            for cell in cells:
                lv_val = cell.strip().upper()
                if re.match(r'^[A-Z0-9]{10,25}$', lv_val) and re.search(r'\d', lv_val):
                    dati["codice_spedizione"] = lv_val; break
            intestazione_lv = False; continue
        if not dati["codice_spedizione"]:
            m = re.search(r'LV\s*[:\-]\s*([A-Z0-9]{10,22})', testo_riga, re.IGNORECASE)
            if m: dati["codice_spedizione"] = m.group(1).strip().upper()

    in_soggetti = False
    for riga in righe:
        cells = [str(c).strip() if c is not None else "" for c in riga]
        testo_riga = " ".join(cells).strip().upper()
        if "CLIENTE/MITTENTE/DESTINATARIO" in testo_riga:
            in_soggetti = True; continue
        if in_soggetti:
            soggetto = cells[0].upper() if cells else ""
            if soggetto in ("MITTENTE","DESTINATARIO"):
                nome = cells[1].strip().upper() if len(cells)>1 else ""
                ind_raw = cells[4] if len(cells)>4 else ""
                tel = cells[5].strip() if len(cells)>5 else ""
                ind, cap_p, citta_p, prov_p = parse_indirizzo(ind_raw)
                cap_citta = f"{cap_p} {citta_p} ({prov_p})" if prov_p else f"{cap_p} {citta_p}"
                if soggetto == "MITTENTE":
                    dati["mit_nome"]=nome; dati["mit_indirizzo"]=ind.upper()
                    dati["mit_cap_citta"]=cap_citta.strip().upper(); dati["mit_tel"]=tel
                else:
                    dati["dest_nome"]=nome; dati["dest_indirizzo"]=ind.upper()
                    dati["dest_cap_citta"]=cap_citta.strip().upper(); dati["dest_tel"]=tel
                    dati["cap"]=cap_p
            if soggetto=="" and dati["dest_nome"] and dati["mit_nome"]:
                in_soggetti=False

    in_peso = False
    for riga in righe:
        cells = [str(c).strip() if c is not None else "" for c in riga]
        testo_riga = " ".join(cells).strip().upper()
        if "PESO E DIMENSIONI" in testo_riga: in_peso=True; continue
        if in_peso and cells and cells[0].upper()=="DICHIARATO":
            try:
                n_colli = cells[3].strip() if len(cells)>3 else ""
                altezza = cells[4].strip().replace(",",".") if len(cells)>4 else ""
                larghez = cells[5].strip().replace(",",".") if len(cells)>5 else ""
                profond = cells[6].strip().replace(",",".") if len(cells)>6 else ""
                peso_val = re.sub(r'[^\d.]', '', cells[7].strip().replace(",",".") if len(cells)>7 else "")
                dati["colli"] = n_colli if n_colli and n_colli!="-" else "1"
                dati["peso"]  = peso_val
                if all(v and v!="-" for v in [altezza,larghez,profond]):
                    def fmt(v):
                        try: f=float(v); return str(int(f)) if f==int(f) else str(f)
                        except: return v
                    dati["dimensioni"] = f"{fmt(altezza)}x{fmt(larghez)}x{fmt(profond)}"
            except: pass
            in_peso=False
    return dati


# ════════════════════════════════════════════════════════════════════════════════
# ─── UI STREAMLIT ─────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3 = st.tabs(["📦 Spedizione", "🔲 Solo Barcode", "📥 Importa da Excel"])

# ── Inizializza session state ──────────────────────────────────────────────────
defaults = {
    "codice_spedizione": "3UW1P11001989",
    "cap": "51039",
    "mit_nome": "CAFFETTERIA STAZIONE",
    "mit_indirizzo": "PIAZZALE XX SETTEMBRE 24",
    "mit_cap_citta": "60044 FABRIANO (AN)",
    "mit_tel": "3761538455",
    "dest_nome": "PIERO VIANI",
    "dest_indirizzo": "VIA TORQUATO TASSO 114",
    "dest_cap_citta": "51039 QUARRATA (PT)",
    "dest_tel": "3924545698",
    "peso": "8.0",
    "colli": "1/1",
    "dimensioni": "40x15x40",
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ════════════════════════════════════════════════════════════════════════════════
# TAB 1 — SPEDIZIONE COMPLETA
# ════════════════════════════════════════════════════════════════════════════════
with tab1:
    col_sx, col_dx, col_acc = st.columns([2, 2, 1.2])

    with col_sx:
        st.markdown('<p class="section-label">📮 Spedizione</p>', unsafe_allow_html=True)
        st.session_state["codice_spedizione"] = st.text_input(
            "Codice spedizione *",
            value=st.session_state["codice_spedizione"],
            max_chars=25
        ).upper()
        st.session_state["cap"] = st.text_input(
            "CAP destinazione *",
            value=st.session_state["cap"],
            max_chars=5
        )
        st.divider()
        st.markdown('<p class="section-label">📤 Mittente</p>', unsafe_allow_html=True)
        for key, lbl in [
            ("mit_nome",      "Nome / Ragione sociale"),
            ("mit_indirizzo", "Indirizzo"),
            ("mit_cap_citta", "CAP e Città"),
            ("mit_tel",       "Telefono"),
        ]:
            st.session_state[key] = st.text_input(lbl, value=st.session_state[key], key=f"in_{key}")

    with col_dx:
        st.markdown('<p class="section-label">📬 Destinatario</p>', unsafe_allow_html=True)
        for key, lbl in [
            ("dest_nome",      "Nome"),
            ("dest_indirizzo", "Indirizzo"),
            ("dest_cap_citta", "CAP e Città"),
            ("dest_tel",       "Telefono"),
        ]:
            st.session_state[key] = st.text_input(lbl, value=st.session_state[key], key=f"in_{key}")
        st.divider()
        st.markdown('<p class="section-label">📦 Pacco</p>', unsafe_allow_html=True)
        for key, lbl in [
            ("peso",       "Peso (kg)"),
            ("colli",      "Colli"),
            ("dimensioni", "Dimensioni HxLxP (cm)"),
        ]:
            st.session_state[key] = st.text_input(lbl, value=st.session_state[key], key=f"in_{key}")

    with col_acc:
        st.markdown('<p class="section-label">⚙️ Accessori</p>', unsafe_allow_html=True)
        ACCESSORI = [
            ("Nessuno", False),
            ("RTZ",     True),
            ("FMP",     True),
            ("APP",     False),
            ("CPR",     True),
            ("CGS",     True),
        ]
        acc_sel = st.radio(
            "Accessorio servizio",
            options=[a[0] for a in ACCESSORI],
            index=0,
            label_visibility="collapsed"
        )
        ha_extra_code = next(a[1] for a in ACCESSORI if a[0] == acc_sel)
        extra_code = ""
        if ha_extra_code:
            extra_code = st.text_input(
                "Codice riferimento (6 char)",
                max_chars=6,
                placeholder="es. A1B2C3"
            ).upper()

    st.divider()

    # Validazione
    codice = st.session_state["codice_spedizione"].strip()
    cap    = st.session_state["cap"].strip()
    errori = []
    if not codice: errori.append("Codice spedizione obbligatorio")
    elif len(codice) not in (13,16,18,22,25): errori.append(f"Lunghezza codice non standard ({len(codice)} char — attesi 13/16/18/22/25)")
    if not cap or len(cap)!=5 or not cap.isdigit(): errori.append("CAP deve essere 5 cifre")

    if errori:
        for e in errori: st.warning(f"⚠️ {e}")
    else:
        acc_val = "" if acc_sel == "Nessuno" else acc_sel
        if st.button("⚡ Genera Etichetta PDF", type="primary", use_container_width=True):
            with st.spinner("Generazione etichetta..."):
                try:
                    tmp_path = tempfile.mktemp(suffix=".pdf")
                    genera_pdf(dict(st.session_state), tmp_path, accessorio=acc_val, extra_code=extra_code)
                    with open(tmp_path,"rb") as f: pdf_bytes = f.read()
                    os.unlink(tmp_path)
                    st.success(f"✅ Etichetta generata — {codice}")
                    st.download_button(
                        "⬇️ Scarica PDF Etichetta",
                        data=pdf_bytes,
                        file_name=f"etichetta_{codice}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Errore generazione: {e}")

# ════════════════════════════════════════════════════════════════════════════════
# TAB 2 — SOLO BARCODE
# ════════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("Genera un'etichetta con i soli barcode (LDV + CAP), senza dati mittente/destinatario.")
    c1, c2 = st.columns(2)
    with c1:
        bc_codice = st.text_input("Codice spedizione *", value=st.session_state["codice_spedizione"], key="bc_codice").upper()
        bc_cap    = st.text_input("CAP *", value=st.session_state["cap"], key="bc_cap", max_chars=5)
    with c2:
        bc_ACCESSORI = ["Nessuno","RTZ","FMP","APP","CPR","CGS"]
        bc_acc = st.selectbox("Accessorio", bc_ACCESSORI, key="bc_acc")
        bc_extra = ""
        if bc_acc not in ("Nessuno","APP"):
            bc_extra = st.text_input("Codice riferimento (6 char)", max_chars=6, key="bc_extra").upper()

    if st.button("⚡ Genera Barcode PDF", type="primary", use_container_width=True, key="btn_bc"):
        if not bc_codice or not bc_cap or len(bc_cap)!=5:
            st.warning("⚠️ Inserisci un codice spedizione e un CAP valido (5 cifre).")
        else:
            with st.spinner("Generazione barcode..."):
                try:
                    tmp_path = tempfile.mktemp(suffix=".pdf")
                    bc_acc_val = "" if bc_acc=="Nessuno" else bc_acc
                    genera_pdf_solo_barcode(bc_codice, bc_cap, tmp_path,
                                            accessorio=bc_acc_val, extra_code=bc_extra)
                    with open(tmp_path,"rb") as f: pdf_bytes = f.read()
                    os.unlink(tmp_path)
                    st.success(f"✅ Barcode generato — {bc_codice}")
                    st.download_button(
                        "⬇️ Scarica PDF Barcode",
                        data=pdf_bytes,
                        file_name=f"barcode_{bc_codice}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Errore: {e}")

# ════════════════════════════════════════════════════════════════════════════════
# TAB 3 — IMPORTA DA EXCEL
# ════════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown("Carica un export Excel di **OneTracking** per compilare automaticamente tutti i campi.")
    uploaded_xls = st.file_uploader("Seleziona file Excel (.xls / .xlsx)", type=["xls","xlsx","xlsm"])
    if uploaded_xls:
        if st.button("📥 Importa dati", type="primary", use_container_width=True):
            with st.spinner("Lettura file Excel..."):
                try:
                    ext = os.path.splitext(uploaded_xls.name)[1].lower()
                    dati_importati = importa_da_excel(uploaded_xls.read(), ext)
                    # Aggiorna session state
                    for k, v in dati_importati.items():
                        if v:
                            st.session_state[k] = v
                    st.success(
                        f"✅ Dati importati!\n\n"
                        f"**Mittente:** {dati_importati.get('mit_nome','-')}\n\n"
                        f"**Destinatario:** {dati_importati.get('dest_nome','-')}\n\n"
                        f"**LV:** {dati_importati.get('codice_spedizione','-')}\n\n"
                        f"**Peso:** {dati_importati.get('peso','-')} kg  |  "
                        f"**Colli:** {dati_importati.get('colli','-')}  |  "
                        f"**Dim:** {dati_importati.get('dimensioni','-')}"
                    )
                    st.info("💡 Vai alla scheda **Spedizione** per verificare i dati e generare l'etichetta.")
                except Exception as e:
                    st.error(f"Errore importazione: {e}")
